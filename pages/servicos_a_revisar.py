import streamlit as st
import pandas as pd
import datetime
import os
import re
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.express as px
import plotly.graph_objects as go
from python_graphs_CIN.utils.data_utils import EXCEL_FILE, load_excel, process_sheet_data, save_excel, SHEET_CONFIG

ARQUIVO = os.path.join(os.path.dirname(__file__), '..', 'revisarservicos.txt')

def parse_data_linha(linha):
    try:
        if linha.strip().lower().startswith('serviço | orgao | tempo | data ultima publicacao'):
            return None

        partes = [p.strip() for p in linha.split('|')]
        servico = ''
        orgao = ''
        tempo = 0
        data_formatada = ''

        if len(partes) >= 1:
            servico = partes[0] if partes[0] else ''
        if len(partes) >= 2:
            segundo_campo = partes[1]
            match_tempo = re.search(r'(\d+)$', segundo_campo)
            if match_tempo:
                tempo = int(match_tempo.group(1))
                orgao = segundo_campo[:match_tempo.start()].strip()
            else:
                orgao = segundo_campo
                tempo = 0
        if len(partes) >= 3:
            try:
                tempo = int(re.findall(r'\d+', partes[2])[0]) if re.findall(r'\d+', partes[2]) else tempo
            except (ValueError, IndexError):
                pass
        if len(partes) >= 4 and partes[3]:
            data_raw = partes[3]
            match = re.match(r'(\d{1,2})\s*de\s*(\w+)\s*(?:de\s*)?(\d{4})(?:\s*às\s*(\d{2}:\d{2}))?', data_raw, re.IGNORECASE)
            if match:
                groups = match.groups()
                dia = groups[0]
                mes = groups[1]
                ano = groups[2]
                hora = groups[3] if len(groups) > 3 else None
                meses = {
                    'january': '01', 'february': '02', 'march': '03', 'april': '04',
                    'may': '05', 'june': '06', 'july': '07', 'august': '08',
                    'september': '09', 'october': '10', 'november': '11', 'december': '12',
                    'janeiro': '01', 'fevereiro': '02', 'março': '03', 'abril': '04',
                    'maio': '05', 'junho': '06', 'julho': '07', 'agosto': '08',
                    'setembro': '09', 'outubro': '10', 'novembro': '11', 'dezembro': '12'
                }
                mes_num = meses.get(mes.lower(), '01')
                data_formatada = f"{dia.zfill(2)}/{mes_num}/{ano}" + (f" {hora}" if hora else "")
            else:
                for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y', '%Y-%m-%d']:
                    try:
                        data_dt = pd.to_datetime(data_raw, format=fmt, errors='coerce')
                        if pd.notna(data_dt):
                            data_formatada = data_dt.strftime('%d/%m/%Y')
                            break
                    except:
                        continue
                else:
                    st.warning(f"Formato de data inválido na linha: {linha}")
                    data_formatada = ''

        return {
            'SERVIÇO': servico,
            'ORGAO': orgao,
            'TEMPO': tempo,
            'DATA ULTIMA PUBLICAÇÃO': data_formatada
        }
    except Exception as e:
        st.error(f"Erro ao processar linha: {linha}. Detalhes: {str(e)}")
        return None

def carregar_dados():
    try:
        with open(ARQUIVO, encoding='utf-8') as f:
            linhas = f.readlines()
        dados = []
        for linha in linhas:
            if linha.strip():
                item = parse_data_linha(linha)
                if item:
                    dados.append(item)
        df = pd.DataFrame(dados)
        if df.empty:
            st.warning("Nenhum dado válido encontrado no arquivo revisarservicos.txt.")
        return df
    except Exception as e:
        st.error(f"Erro ao carregar o arquivo revisarservicos.txt: {str(e)}")
        return pd.DataFrame()

def atualizar_tempo(df):
    hoje = datetime.datetime.now()
    for idx, row in df.iterrows():
        data_str = row['DATA ULTIMA PUBLICAÇÃO']
        if data_str and data_str.strip():
            try:
                for fmt in ['%d/%m/%Y %H:%M', '%d/%m/%Y']:
                    try:
                        data_dt = datetime.datetime.strptime(data_str, fmt)
                        dias = (hoje.date() - data_dt.date()).days
                        df.at[idx, 'TEMPO'] = dias
                        break
                    except ValueError:
                        continue
                else:
                    df.at[idx, 'TEMPO'] = 0
                    st.warning(f"Data inválida na linha {idx}: {data_str}")
            except Exception as e:
                df.at[idx, 'TEMPO'] = 0
                st.warning(f"Erro ao processar data na linha {idx}: {data_str}. Detalhes: {str(e)}")
        else:
            df.at[idx, 'TEMPO'] = 0
    return df

def salvar_dados(df):
    with open(ARQUIVO, 'w', encoding='utf-8') as f:
        f.write('SERVIÇO | ORGAO | TEMPO | DATA ULTIMA PUBLICACAO\n')
        for _, row in df.iterrows():
            data_str = row['DATA ULTIMA PUBLICAÇÃO'] if row['DATA ULTIMA PUBLICAÇÃO'] else ''
            f.write(f"{row['SERVIÇO']} | {row['ORGAO']} | {row['TEMPO']} | {data_str}\n")

def to_excel(df):
    output = BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Serviços a Revisar'

    headers = df.columns.tolist()
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for row in dataframe_to_rows(df, index=False, header=False):
        worksheet.append(row)

    for col in range(1, len(df.columns) + 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    workbook.save(output)
    return output.getvalue()

def generate_servicos_a_revisar_dashboards(df, limite_grafico="Sem Limites"):
    """Gera gráficos para a aba 'Serviços a Revisar'."""
    if df.empty:
        return []
    
    df_plot = df.head(limite_grafico) if isinstance(limite_grafico, int) else df
    figs = []
    
    orgao_counts = df_plot['ORGAO'].value_counts().reset_index()
    orgao_counts.columns = ['Órgão', 'Quantidade']
    fig1 = px.bar(
        orgao_counts,
        x='Órgão',
        y='Quantidade',
        title='Distribuição de Serviços por Órgão',
        text='Quantidade',
        color='Órgão',
        color_discrete_sequence=px.colors.qualitative.Plotly
    )
    fig1.update_traces(textposition='outside')
    fig1.update_layout(showlegend=False, xaxis_title="Órgão", yaxis_title="Número de Serviços")
    figs.append(fig1)
    
    df_tempo = df_plot[df_plot['TEMPO'] > 0]
    if not df_tempo.empty:
        fig2 = px.histogram(
            df_tempo,
            x='TEMPO',
            title='Distribuição do Tempo desde a Última Publicação (Dias)',
            nbins=20,
            color_discrete_sequence=['#4F81BD']
        )
        fig2.update_layout(xaxis_title="Dias desde a Última Publicação", yaxis_title="Contagem")
        figs.append(fig2)
    
    df_data = df_plot[df_plot['DATA ULTIMA PUBLICAÇÃO'].notna() & (df_plot['DATA ULTIMA PUBLICAÇÃO'] != '')]
    if not df_data.empty:
        df_data['DATA'] = pd.to_datetime(df_data['DATA ULTIMA PUBLICAÇÃO'], format='%d/%m/%Y', errors='coerce')
        df_data = df_data[df_data['DATA'].notna()]
        df_data['Mês-Ano'] = df_data['DATA'].dt.strftime('%Y-%m')
        data_counts = df_data['Mês-Ano'].value_counts().sort_index().reset_index()
        data_counts.columns = ['Mês-Ano', 'Quantidade']
        fig3 = px.line(
            data_counts,
            x='Mês-Ano',
            y='Quantidade',
            title='Serviços por Mês de Última Publicação',
            markers=True,
            text='Quantidade'
        )
        fig3.update_traces(mode='lines+markers+text', textposition='top center')
        fig3.update_layout(xaxis_title="Mês-Ano", yaxis_title="Número de Serviços")
        figs.append(fig3)
    
    return figs

def render_servicos_a_revisar():
    st.markdown("""
        <h3>Serviços a Revisar <span class="material-icons" style="vertical-align: middle; color: #004aad;">checklist</span></h3>
    """, unsafe_allow_html=True)
    
    df = carregar_dados()
    if df.empty:
        st.error("Nenhum dado carregado do arquivo revisarservicos.txt.")
        st.markdown("""
        ### Possíveis Soluções
        - Verifique se o arquivo `revisarservicos.txt` está no diretório raiz do projeto.
        - Confirme se o arquivo contém dados no formato correto (ex.: SERVIÇO | ORGAO | TEMPO | DATA ULTIMA PUBLICACAO).
        - Exemplo de linha válida: 'Serviço A | Cidade X | 150 | 15 de Janeiro de 2025 às 10:00'
        """)
        return
    
    df = atualizar_tempo(df)
    tab1, tab2 = st.tabs(["Dados", "Dashboard"])

    with tab1:
        st.markdown("### Dados de Serviços a Revisar")
        limite_listagem = st.selectbox("Limitar número de serviços na listagem", [2, 5, 10, 15, 20, 50, 100, "Sem Limites"], index=7, key="limit_list")
        df_display = df.head(limite_listagem) if limite_listagem != "Sem Limites" else df
        st.dataframe(df_display, use_container_width=True)

        excel_data = to_excel(df_display)
        st.download_button(
            label="Exportar como Excel",
            data=excel_data,
            file_name="servicos_a_revisar.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        with st.expander("Adicionar Novo Serviço"):
            servico = st.text_input("Serviço", key="new_servico")
            orgao = st.text_input("Órgão", key="new_orgao")
            data_pub = st.date_input("Data Última Publicação (opcional)", value=None, key="new_data")
            hora_pub = st.time_input("Hora Última Publicação (opcional)", value=None, key="new_hora") if data_pub else None
            if st.button("Adicionar Serviço", key="add_button"):
                if servico and orgao:
                    nova_data = f"{data_pub.strftime('%d/%m/%Y') if data_pub else ''} {hora_pub.strftime('%H:%M') if hora_pub else ''}".strip() if data_pub and hora_pub else ''
                    novo = {'SERVIÇO': servico, 'ORGAO': orgao, 'TEMPO': 0, 'DATA ULTIMA PUBLICAÇÃO': nova_data}
                    df = pd.concat([df, pd.DataFrame([novo])], ignore_index=True)
                    df = atualizar_tempo(df)
                    salvar_dados(df)
                    st.success("Serviço adicionado!")
                    st.rerun()
                else:
                    st.error("Serviço e Órgão são obrigatórios.")

        with st.expander("Editar ou Apagar Serviço"):
            if not df.empty:
                idx = st.selectbox("Selecione uma linha", df.index, key="edit_idx")
                row = df.loc[idx]
                with st.expander("Editar Serviço"):
                    servico_edit = st.text_input("Serviço", value=row['SERVIÇO'], key="edit_servico")
                    orgao_edit = st.text_input("Órgão", value=row['ORGAO'], key="edit_orgao")
                    data_edit = st.date_input(
                        "Data Última Publicação (opcional)",
                        value=pd.to_datetime(row['DATA ULTIMA PUBLICAÇÃO'], errors='coerce').date() if pd.notna(row['DATA ULTIMA PUBLICAÇÃO']) else None,
                        key="edit_data"
                    )
                    hora_edit = st.time_input(
                        "Hora Última Publicação (opcional)",
                        value=pd.to_datetime(row['DATA ULTIMA PUBLICAÇÃO'], errors='coerce').time() if pd.notna(row['DATA ULTIMA PUBLICAÇÃO']) and ' ' in row['DATA ULTIMA PUBLICAÇÃO'] else None,
                        key="edit_hora"
                    ) if data_edit else None
                    if st.button("Salvar Edição", key="save_button"):
                        if servico_edit and orgao_edit:
                            nova_data = f"{data_edit.strftime('%d/%m/%Y') if data_edit else ''} {hora_edit.strftime('%H:%M') if hora_edit else ''}".strip() if data_edit and hora_edit else ''
                            df.at[idx, 'SERVIÇO'] = servico_edit
                            df.at[idx, 'ORGAO'] = orgao_edit
                            df.at[idx, 'DATA ULTIMA PUBLICAÇÃO'] = nova_data
                            df = atualizar_tempo(df)
                            salvar_dados(df)
                            st.success("Serviço atualizado!")
                            st.rerun()
                        else:
                            st.error("Serviço e Órgão são obrigatórios.")
                with st.expander("Apagar Serviço"):
                    if st.button("Confirmar Apagar", key="delete_button"):
                        df = df.drop(idx).reset_index(drop=True)
                        salvar_dados(df)
                        st.success("Serviço removido!")
                        st.rerun()
            else:
                st.warning("Nenhuma linha disponível para editar ou apagar.")

    with tab2:
        st.markdown("### Dashboard de Serviços a Revisar")
        servico_selecionado = st.selectbox("Selecione um Serviço (opcional)", ["Nenhum"] + df['SERVIÇO'].tolist(), key="servico_select")
        limite_grafico = st.selectbox("Limitar número de serviços nos gráficos", [2, 5, 10, 15, 20, 50, 100, "Sem Limites"], index=7, key="limit_graph")
        figs = generate_servicos_a_revisar_dashboards(df, limite_grafico)
        for fig in figs:
            st.plotly_chart(fig, use_container_width=True)

if __name__ == "__main__":
    render_servicos_a_revisar()
import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.express as px
from python_graphs_CIN.utils.data_utils import load_excel, process_sheet_data, save_excel, SHEET_CONFIG, EXCEL_FILE

@st.cache_data
def load_and_process_ag_instalacao(_file_path=EXCEL_FILE):
    """Carrega e processa a aba Ag_Instalacao com caching."""
    try:
        raw_df = load_excel('Ag_Instalacao', _file_path)
        if raw_df.empty:
            st.error("Nenhum dado disponível para a aba 'Ag_Instalacao'. Verifique o nome da aba no arquivo Excel.")
            return pd.DataFrame()
        
        df = process_sheet_data(raw_df, 'Ag_Instalacao')
        return df
    except Exception as e:
        st.error(f"Erro ao processar a aba Ag_Instalacao: {str(e)}")
        return pd.DataFrame()

def to_excel(df):
    """Converte DataFrame para Excel com estilização."""
    output = BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Relatório de Datas'

    headers = df.columns.tolist()
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for row in dataframe_to_rows(df, index=False, header=False):
        worksheet.append(row)

    for col in range(1, len(df.columns) + 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    workbook.save(output)
    return output.getvalue()

def generate_ag_instalacao_dashboards(df, limite_cidades="Sem Limites"):
    """Gera gráficos e relatório para a aba 'Ag_Instalacao'."""
    if df.empty:
        return [], pd.DataFrame()
    
    # Aplicar limite ao número de cidades, se necessário
    df_plot = df.head(limite_cidades) if isinstance(limite_cidades, int) else df
    
    figs = []
    
    # Gráfico 1: Distribuição por Situação da Infra-estrutura
    sit_infra_counts = df_plot['SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA'].value_counts().reset_index()
    sit_infra_counts.columns = ['Situação', 'Quantidade']
    fig1 = px.pie(
        sit_infra_counts,
        values='Quantidade',
        names='Situação',
        title='Distribuição por Situação da Infra-estrutura',
        color_discrete_sequence=px.colors.qualitative.Plotly
    )
    fig1.update_traces(textinfo='percent+label')
    figs.append(fig1)
    
    # Gráfico 2: Distribuição por Parecer da Visita Técnica
    parecer_counts = df_plot['PARECER DA VISITA TÉCNICA'].value_counts().reset_index()
    parecer_counts.columns = ['Parecer', 'Quantidade']
    fig2 = px.bar(
        parecer_counts,
        x='Parecer',
        y='Quantidade',
        title='Distribuição por Parecer da Visita Técnica',
        text='Quantidade',
        color='Parecer',
        color_discrete_sequence=px.colors.qualitative.Plotly
    )
    fig2.update_traces(textposition='outside')
    fig2.update_layout(showlegend=False)
    figs.append(fig2)
    
    # Gráfico 3: Status do Termo de Cooperação
    termo_counts = df_plot['SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO'].value_counts().reset_index()
    termo_counts.columns = ['Situação', 'Quantidade']
    fig3 = px.bar(
        termo_counts,
        x='Situação',
        y='Quantidade',
        title='Distribuição por Situação do Termo de Cooperação',
        text='Quantidade',
        color='Situação',
        color_discrete_sequence=px.colors.qualitative.Plotly
    )
    fig3.update_traces(textposition='outside')
    fig3.update_layout(showlegend=False)
    figs.append(fig3)
    
    # Relatório de datas
    df_relatorio = df_plot[['CIDADE', 'DATA DO D.O.', 'DATA DA INSTALAÇÃO', 'DATA DO INÍCIO ATEND.']].copy()
    df_relatorio['DATA DO D.O.'] = pd.to_datetime(df_relatorio['DATA DO D.O.'], format='%d/%m/%Y', errors='coerce')
    df_relatorio['DATA DA INSTALAÇÃO'] = pd.to_datetime(df_relatorio['DATA DA INSTALAÇÃO'], format='%d/%m/%Y', errors='coerce')
    df_relatorio['DATA DO INÍCIO ATEND.'] = pd.to_datetime(df_relatorio['DATA DO INÍCIO ATEND.'], format='%d/%m/%Y', errors='coerce')
    
    return figs, df_relatorio

def render_ag_instalacao(uploaded_file=None):
    st.markdown("""
        <h3>Aguardando Instalação <span class="material-icons" style="vertical-align: middle; color: #004aad;">construction</span></h3>
    """, unsafe_allow_html=True)
    
    file_path = uploaded_file if uploaded_file else EXCEL_FILE
    df = load_and_process_ag_instalacao(file_path)
    
    if df.empty:
        st.error("Nenhum dado disponível para a aba 'Ag_Instalacao'. Verifique o arquivo Excel e o nome da aba.")
        return
    
    st.markdown("### Tabela Completa")
    st.dataframe(df, use_container_width=True)
    
    # Aba de navegação: Dados e Dashboard
    tab1, tab2 = st.tabs(["Dados", "Dashboard"])
    
    with tab1:
        # Adicionar Novo Registro
        with st.expander("Adicionar Novo Registro"):
            cidade = st.text_input("Cidade", key="new_cidade")
            sit_infra = st.selectbox(
                "Situação da Infra-estrutura p/ Visita Técnica",
                SHEET_CONFIG['Ag_Instalacao']['columns']['SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA']['values'],
                key="new_sit_infra"
            )
            parecer_visita = st.selectbox(
                "Parecer da Visita Técnica",
                SHEET_CONFIG['Ag_Instalacao']['columns']['PARECER DA VISITA TÉCNICA']['values'],
                key="new_parecer_visita"
            )
            realizou_treinamento = st.checkbox("Realizou Treinamento?", key="new_realizou_treinamento")
            situacao_termo = st.selectbox(
                "Situação do Novo Termo de Cooperação",
                SHEET_CONFIG['Ag_Instalacao']['columns']['SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO']['values'],
                key="new_situacao_termo"
            )
            data_do = st.date_input("Data do D.O. (opcional)", value=None, key="new_data_do")
            apto_instalacao = st.checkbox("Apto para Instalação?", key="new_apto_instalacao")
            data_instalacao = st.date_input("Data da Instalação (opcional)", value=None, key="new_data_instalacao")
            data_inicio_atend = st.date_input("Data do Início Atend. (opcional)", value=None, key="new_data_inicio_atend")
            
            if st.button("Adicionar Registro", key="add_button"):
                if cidade:
                    novo = {
                        'CIDADE': cidade,
                        'SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA': sit_infra,
                        'PARECER DA VISITA TÉCNICA': parecer_visita,
                        'REALIZOU TREINAMENTO?': realizou_treinamento,
                        'SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO': situacao_termo,
                        'DATA DO D.O.': data_do.strftime('%d/%m/%Y') if data_do else '',
                        'APTO PARA INSTALAÇÃO': apto_instalacao,
                        'DATA DA INSTALAÇÃO': data_instalacao.strftime('%d/%m/%Y') if data_instalacao else '',
                        'DATA DO INÍCIO ATEND.': data_inicio_atend.strftime('%d/%m/%Y') if data_inicio_atend else ''
                    }
                    df = pd.concat([df, pd.DataFrame([novo])], ignore_index=True)
                    if save_excel(df, 'Ag_Instalacao', file_path):
                        st.cache_data.clear()
                        st.success("Registro adicionado!")
                        st.rerun()
                else:
                    st.error("Cidade é obrigatória.")
        
        # Editar ou Apagar Registro
        with st.expander("Editar ou Apagar Registro"):
            if not df.empty:
                idx = st.selectbox("Selecione uma linha", df.index, key="edit_idx")
                row = df.loc[idx]
                
                with st.expander("Editar Registro"):
                    cidade_edit = st.text_input("Cidade", value=row['CIDADE'], key="edit_cidade")
                    sit_infra_edit = st.selectbox(
                        "Situação da Infra-estrutura p/ Visita Técnica",
                        SHEET_CONFIG['Ag_Instalacao']['columns']['SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA']['values'],
                        index=SHEET_CONFIG['Ag_Instalacao']['columns']['SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA']['values'].index(row['SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA']) if row['SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA'] in SHEET_CONFIG['Ag_Instalacao']['columns']['SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA']['values'] else 0,
                        key="edit_sit_infra"
                    )
                    parecer_visita_edit = st.selectbox(
                        "Parecer da Visita Técnica",
                        SHEET_CONFIG['Ag_Instalacao']['columns']['PARECER DA VISITA TÉCNICA']['values'],
                        index=SHEET_CONFIG['Ag_Instalacao']['columns']['PARECER DA VISITA TÉCNICA']['values'].index(row['PARECER DA VISITA TÉCNICA']) if row['PARECER DA VISITA TÉCNICA'] in SHEET_CONFIG['Ag_Instalacao']['columns']['PARECER DA VISITA TÉCNICA']['values'] else 0,
                        key="edit_parecer_visita"
                    )
                    realizou_treinamento_edit = st.checkbox(
                        "Realizou Treinamento?",
                        value=row['REALIZOU TREINAMENTO?'],
                        key="edit_realizou_treinamento"
                    )
                    situacao_termo_edit = st.selectbox(
                        "Situação do Novo Termo de Cooperação",
                        SHEET_CONFIG['Ag_Instalacao']['columns']['SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO']['values'],
                        index=SHEET_CONFIG['Ag_Instalacao']['columns']['SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO']['values'].index(row['SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO']) if row['SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO'] in SHEET_CONFIG['Ag_Instalacao']['columns']['SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO']['values'] else 0,
                        key="edit_situacao_termo"
                    )
                    data_do_edit = st.date_input(
                        "Data do D.O. (opcional)",
                        value=pd.to_datetime(row['DATA DO D.O.'], format='%d/%m/%Y', errors='coerce').date() if pd.notna(row['DATA DO D.O.']) and row['DATA DO D.O.'] else None,
                        key="edit_data_do"
                    )
                    apto_instalacao_edit = st.checkbox(
                        "Apto para Instalação?",
                        value=row['APTO PARA INSTALAÇÃO'],
                        key="edit_apto_instalacao"
                    )
                    data_instalacao_edit = st.date_input(
                        "Data da Instalação (opcional)",
                        value=pd.to_datetime(row['DATA DA INSTALAÇÃO'], format='%d/%m/%Y', errors='coerce').date() if pd.notna(row['DATA DA INSTALAÇÃO']) and row['DATA DA INSTALAÇÃO'] else None,
                        key="edit_data_instalacao"
                    )
                    data_inicio_atend_edit = st.date_input(
                        "Data do Início Atend. (opcional)",
                        value=pd.to_datetime(row['DATA DO INÍCIO ATEND.'], format='%d/%m/%Y', errors='coerce').date() if pd.notna(row['DATA DO INÍCIO ATEND.']) and row['DATA DO INÍCIO ATEND.'] else None,
                        key="edit_data_inicio_atend"
                    )
                    
                    if st.button("Salvar Edição", key="save_button"):
                        if cidade_edit:
                            df.at[idx, 'CIDADE'] = cidade_edit
                            df.at[idx, 'SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA'] = sit_infra_edit
                            df.at[idx, 'PARECER DA VISITA TÉCNICA'] = parecer_visita_edit
                            df.at[idx, 'REALIZOU TREINAMENTO?'] = realizou_treinamento_edit
                            df.at[idx, 'SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO'] = situacao_termo_edit
                            df.at[idx, 'DATA DO D.O.'] = data_do_edit.strftime('%d/%m/%Y') if data_do_edit else ''
                            df.at[idx, 'APTO PARA INSTALAÇÃO'] = apto_instalacao_edit
                            df.at[idx, 'DATA DA INSTALAÇÃO'] = data_instalacao_edit.strftime('%d/%m/%Y') if data_instalacao_edit else ''
                            df.at[idx, 'DATA DO INÍCIO ATEND.'] = data_inicio_atend_edit.strftime('%d/%m/%Y') if data_inicio_atend_edit else ''
                            if save_excel(df, 'Ag_Instalacao', file_path):
                                st.cache_data.clear()
                                st.success("Registro atualizado!")
                                st.rerun()
                        else:
                            st.error("Cidade é obrigatória.")
                
                with st.expander("Apagar Registro"):
                    if st.button("Confirmar Apagar", key="delete_button"):
                        df = df.drop(idx).reset_index(drop=True)
                        if save_excel(df, 'Ag_Instalacao', file_path):
                            st.cache_data.clear()
                            st.success("Registro removido!")
                            st.rerun()
            else:
                st.warning("Nenhuma linha disponível para editar ou apagar.")
    
    with tab2:
        st.markdown("### Dashboard de Aguardando Instalação")
        
        limite_cidades = st.selectbox(
            "Limitar número de cidades no dashboard",
            [2, 5, 10, 15, 20, 50, 100, "Sem Limites"],
            index=7,
            key="limit_cidades"
        )
        
        figs, df_relatorio = generate_ag_instalacao_dashboards(df, limite_cidades)
        for fig in figs:
            st.plotly_chart(fig, use_container_width=True)
        
        st.markdown("#### Relatório de Datas por Cidade")
        if not df_relatorio.empty:
            def style_dataframe(df):
                return df.style.set_table_styles([
                    {'selector': 'th', 'props': [('background-color', '#4F81BD'), ('color', 'white'), ('font-weight', 'bold'), ('text-align', 'center'), ('padding', '8px')]},
                    {'selector': 'td', 'props': [('border', '1px solid #ddd'), ('padding', '8px'), ('text-align', 'center')]},
                    {'selector': 'tr:nth-child(even)', 'props': [('background-color', '#f2f2f2')]},
                    {'selector': 'tr:hover', 'props': [('background-color', '#e0e0e0')]}
                ]).set_properties(**{'font-size': '14px'})
            st.dataframe(style_dataframe(df_relatorio), use_container_width=True)
        
        excel_data = to_excel(df_relatorio)
        st.download_button(
            label="Exportar Relatório como Excel",
            data=excel_data,
            file_name="relatorio_datas_instalacao.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if __name__ == "__main__":
    render_ag_instalacao()
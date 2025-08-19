import pandas as pd
import re
from datetime import datetime
import numpy as np
import streamlit as st
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

# Caminho padrão do arquivo Excel
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "..", "ACOMPANHAMENTO_CIN_EM_TODO_LUGAR.xlsx")
FALLBACK_EXCEL_FILE = r"C:/Users/re049227/Documents/python_graphs_CIN/ACOMPANHAMENTO_CIN_EM_TODO_LUGAR.xlsx"

# Normalizar nomes das abas para comparação
def normalize_sheet_name(name):
    """Normaliza nomes de abas removendo acentos, pontos e convertendo para minúsculas."""
    if not name:
        return name
    name = name.replace('.', ' ').replace('_', ' ').replace('-', ' ')
    name = re.sub(r'[áàãâä]', 'a', name.lower())
    name = re.sub(r'[éèêë]', 'e', name)
    name = re.sub(r'[íìîï]', 'i', name)
    name = re.sub(r'[óòõôö]', 'o', name)
    name = re.sub(r'[úùûü]', 'u', name)
    name = re.sub(r'[ç]', 'c', name)
    return name.strip()

# Mapeamento de nomes normalizados para nomes reais
SHEET_NAME_MAPPING = {
    normalize_sheet_name(sheet): sheet for sheet in [
        'Geral-Amplo', 'Lista X', 'Geral-Resumo', 'Status', 'Visitas Realizadas',
        'Ag. Visita', 'Ag_info_prefeitura', 'Ag_Instalacao', 'Publicados', 'Instalados',
        'Funcionando', 'Treina-turma', 'Treina-cidade', 'Informações', 'Chefes_Posto', 'Produtividade'
    ]
}

# Configuração das abas com colunas e tipos
SHEET_CONFIG = {
    'Geral-Amplo': {'columns': {'CIDADE': {'type': 'string'}}},
    'Lista X': {'columns': {'CIDADE': {'type': 'string'}}},
    'Geral-Resumo': {'columns': {'CIDADE': {'type': 'string'}}},
    'Status': {'columns': {'CIDADE': {'type': 'string'}}},
    'Visitas Realizadas': {
        'columns': {
            'CIDADE': {'type': 'string'},
            'DATA DA VISITA': {'type': 'date', 'format': '%d/%m/%Y'},
            'PARECER DA VISITA TÉCNICA': {'type': 'categorical', 'values': ['Aprovado', 'Pendente', 'Reprovado']},
            'APTO PARA INSTALAÇÃO?': {'type': 'boolean'},
            'OBSERVAÇÃO': {'type': 'string'}
        }
    },
    'Ag. Visita': {
        'columns': {
            'CIDADE': {'type': 'string'},
            'SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA': {
                'type': 'categorical',
                'values': ['Sem pendência', 'Com Pendência', 'Não Informada']
            },
            'DATA DA VISITA TÉCNICA': {'type': 'date', 'format': '%d/%m/%Y'},
            'PARECER DA VISITA TÉCNICA': {
                'type': 'categorical',
                'values': ['Reprovado', 'Aprovado', '']
            },
            'ADEQUAÇÕES APÓS VISITA TÉCNICA REALIZADAS': {'type': 'string'},
            'DATA DE FINALIZAÇÃO DAS ADEQUAÇÕES': {'type': 'date', 'format': '%d/%m/%Y'}
        }
    },
    'Ag_info_prefeitura': {
        'columns': {
            'CIDADE': {'type': 'string'},
            'SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA': {
                'type': 'categorical',
                'values': ['Sem pendência', 'Com Pendência', 'Não Informada']
            },
            'DATA DA VISITA TÉCNICA': {'type': 'date', 'format': '%d/%m/%Y'}
        }
    },
    'Ag_Instalacao': {
        'columns': {
            'CIDADE': {'type': 'string'},
            'SIT. DA INFRA-ESTRUTURA P/VISITA TÉCNICA': {
                'type': 'categorical',
                'values': ['Sem pendência', 'Com Pendência', 'Não Informada']
            },
            'PARECER DA VISITA TÉCNICA': {
                'type': 'categorical',
                'values': ['Aprovado', 'Reprovado']
            },
            'REALIZOU TREINAMENTO?': {'type': 'boolean'},
            'SITUAÇÃO DO NOVO TERMO DE COOPERAÇÃO': {
                'type': 'categorical',
                'values': ['Publicado D.O.', 'Não Publicado D.O.']
            },
            'DATA DO D.O.': {'type': 'date', 'format': '%d/%m/%Y'},
            'APTO PARA INSTALAÇÃO': {'type': 'boolean'},
            'DATA DA INSTALAÇÃO': {'type': 'date', 'format': '%d/%m/%Y'},
            'DATA DO INÍCIO ATEND.': {'type': 'date', 'format': '%d/%m/%Y'}
        }
    },
    'Publicados': {'columns': {'CIDADE': {'type': 'string'}}},
    'Instalados': {
        'columns': {
            'CIDADE': {'type': 'string'},
            'DATA DA INSTALAÇÃO': {'type': 'date', 'format': '%d/%m/%Y'},
            'PREFEITURAS DE': {'type': 'string'}
        }
    },
    'Funcionando': {'columns': {'CIDADE': {'type': 'string'}}},
    'Treina-turma': {
        'columns': {
            'CIDADE': {'type': 'string'},
            'PERÍODO PREVISTO DE TREINAMENTO_INÍCIO': {'type': 'date', 'format': '%d/%m/%Y'},
            'PERÍODO PREVISTO DE TREINAMENTO_FIM': {'type': 'date', 'format': '%d/%m/%Y'},
            'TURMA': {'type': 'int'},
            'REALIZOU TREINAMENTO?': {'type': 'boolean'}
        }
    },
    'Treina-cidade': {'columns': {'CIDADE': {'type': 'string'}}},
    'Informações': {'columns': {'CIDADE': {'type': 'string'}}},
    'Chefes_Posto': {
        'columns': {
            'CIDADE': {'type': 'string'},
            'POSTO': {'type': 'string'},
            'NOME': {'type': 'string'},
            'E-MAIL': {'type': 'email'},
            'TELEFONE': {'type': 'phone'},
            'TURMA': {'type': 'int'},
            'DATA TREINAMENTO': {'type': 'training_period'},
            'USUÁRIO': {'type': 'string'}
        }
    },
    'Produtividade': {
        'columns': {
            'CIDADE': {'type': 'string'},
            'PREFEITURAS DE': {'type': 'string'},
            'REALIZOU TREINAMENTO?': {'type': 'boolean'},
            'DATA DA INSTALAÇÃO': {'type': 'date', 'format': '%d/%m/%Y'},
            'DATA DO INÍCIO ATEND.': {'type': 'date', 'format': '%d/%m/%Y'},
            'PERÍODO PREVISTO DE TREINAMENTO_INÍCIO': {'type': 'date', 'format': '%d/%m/%Y'},
            'PERÍODO PREVISTO DE TREINAMENTO_FIM': {'type': 'date', 'format': '%d/%m/%Y'},
            'JANEIRO': {'type': 'float'},
            'FEVEREIRO': {'type': 'float'},
            'MARÇO': {'type': 'float'},
            'ABRIL': {'type': 'float'},
            'MAIO': {'type': 'float'},
            'JUNHO': {'type': 'float'},
            'JULHO': {'type': 'float'},
            'AGOSTO': {'type': 'float'},
            'SETEMBRO': {'type': 'float'},
            'OUTUBRO': {'type': 'float'},
            'NOVEMBRO': {'type': 'float'},
            'DEZEMBRO': {'type': 'float'}
        }
    }
}

@st.cache_data
def load_excel(sheet_name, _file_path=EXCEL_FILE):
    """Carrega uma aba específica do arquivo Excel com caching."""
    try:
        file_path = _file_path
        normalized_sheet_name = normalize_sheet_name(sheet_name)
        actual_sheet_name = SHEET_NAME_MAPPING.get(normalized_sheet_name, sheet_name)
        
        # Handle uploaded file
        if isinstance(file_path, BytesIO):
            st.write(f"[DEBUG] Usando arquivo carregado via upload")
            xls = pd.ExcelFile(file_path, engine='openpyxl')
        else:
            # Check main path
            if not os.path.exists(file_path):
                st.warning(f"Arquivo não encontrado no caminho principal: {os.path.abspath(file_path)}")
                file_path = FALLBACK_EXCEL_FILE
                if not os.path.exists(file_path):
                    st.error(f"Arquivo não encontrado no caminho alternativo: {os.path.abspath(file_path)}")
                    st.markdown("""
                    ### Possíveis Soluções
                    - Verifique se o arquivo `ACOMPANHAMENTO_CIN_EM_TODO_LUGAR.xlsx` está em `C:\\Users\\re049227\\Documents\\python_graphs_CIN\\`.
                    - Confirme se o nome do arquivo está correto (sem espaços extras ou caracteres ocultos).
                    - Se o arquivo foi carregado via upload, certifique-se de que o módulo `upload_excel.py` está configurado corretamente.
                    - Em produção, verifique o caminho do arquivo no servidor ou contêiner.
                    - Certifique-se de que o arquivo não está corrompido ou bloqueado por outro programa.
                    """)
                    return pd.DataFrame()
            xls = pd.ExcelFile(file_path, engine='openpyxl')
        
        # Verify sheet existence
        if actual_sheet_name not in xls.sheet_names:
            st.error(f"Aba '{actual_sheet_name}' não encontrada no arquivo Excel. Abas disponíveis: {xls.sheet_names}")
            return pd.DataFrame()
        
        df = pd.read_excel(file_path, sheet_name=actual_sheet_name, engine='openpyxl')
        df.columns = df.columns.str.replace('\n', ' ').str.strip().str.replace(r'\s+', ' ', regex=True)
        df.columns = df.columns.str.replace('PREFEITURA DE', 'PREFEITURAS DE', regex=False)
        
        expected_cols = list(SHEET_CONFIG.get(actual_sheet_name, {}).get('columns', {}).keys())
        missing_cols = [col for col in expected_cols if col not in df.columns]
        if missing_cols:
            st.warning(f"Colunas ausentes na aba '{actual_sheet_name}': {', '.join(missing_cols)}")
            for col in missing_cols:
                col_type = SHEET_CONFIG.get(actual_sheet_name, {}).get('columns', {}).get(col, {}).get('type', 'string')
                if col_type in ['date', 'datetime']:
                    df[col] = pd.NaT
                elif col_type == 'boolean':
                    df[col] = False
                elif col_type in ['int', 'float']:
                    df[col] = 0
                elif col_type == 'categorical':
                    df[col] = SHEET_CONFIG.get(actual_sheet_name, {}).get('columns', {}).get(col, {}).get('values', [''])[0]
                else:
                    df[col] = ''
        
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(lambda x: str(x).replace('\n', ' ').strip() if pd.notnull(x) else '')
        
        return df
    except Exception as e:
        st.error(f"Erro ao carregar a aba {sheet_name}: {str(e)}")
        st.write(f"[DEBUG] Abas disponíveis no arquivo: {xls.sheet_names if 'xls' in locals() else 'Arquivo não carregado'}")
        return pd.DataFrame()

def parse_training_period(period):
    """Parseia o período de treinamento no formato 'dd/mm a dd/mm/yy' e retorna as datas de início e fim."""
    if pd.isna(period) or period in ['-', '', 'VAZIO', 'N-PREV.', 'nan']:
        return pd.Series([pd.NaT, pd.NaT])
    
    try:
        period = str(period).strip()
        if not period or period in ['-', 'VAZIO', 'N-PREV.']:
            return pd.Series([pd.NaT, pd.NaT])
        
        parts = re.split(r'\s*(?:à|a)\s*', period, flags=re.IGNORECASE)
        if len(parts) != 2:
            return pd.Series([pd.NaT, pd.NaT])
        
        start_date, end_date = parts
        # Adicionar ano à primeira data, se necessário
        if len(start_date.split('/')) == 2:
            start_date = f"{start_date}/20{end_date[-2:]}"
        elif len(start_date.split('/')) == 3 and len(start_date.split('/')[-1]) == 2:
            start_date = f"{start_date[:6]}20{start_date[-2:]}"
        
        if len(end_date.split('/')) == 2:
            end_date = f"{end_date}/20{end_date[-2:]}"
        elif len(end_date.split('/')) == 3 and len(end_date.split('/')[-1]) == 2:
            end_date = f"{end_date[:6]}20{end_date[-2:]}"
        
        for fmt in ['%d/%m/%Y', '%d/%m/%y']:
            try:
                start_dt = pd.to_datetime(start_date, format=fmt, errors='coerce')
                end_dt = pd.to_datetime(end_date, format=fmt, errors='coerce')
                if pd.notna(start_dt) and pd.notna(end_dt):
                    return pd.Series([start_dt, end_dt])
            except:
                continue
        
        return pd.Series([pd.NaT, pd.NaT])
    except Exception:
        return pd.Series([pd.NaT, pd.NaT])

def clean_phone_number(phone):
    """Limpa e padroniza números de telefone."""
    if pd.isna(phone) or phone in ['-', 'sn', 'vazio', 'VAZIO', 'nan']:
        return ''
    phone = str(phone).strip()
    phone = re.sub(r'[^\d]', '', phone)
    if len(phone) >= 10:
        return f"({phone[:2]}) {phone[2:7]}-{phone[7:]}"
    return ''

def clean_email(email):
    """Limpa e padroniza e-mails."""
    if pd.isna(email) or email in ['-', 'sn', 'vazio', 'VAZIO', 'nan']:
        return ''
    email = str(email).strip().lower()
    if re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
        return email
    return ''

def clean_time(time):
    """Limpa e padroniza horários."""
    if pd.isna(time) or time in ['-', 'sn', 'vazio', 'VAZIO', 'nan']:
        return ''
    time = str(time).strip()
    try:
        pd.to_datetime(time, format='%H:%M:%S')
        return time
    except:
        try:
            pd.to_datetime(time, format='%H:%M')
            return time + ':00'
        except:
            return ''

@st.cache_data
def process_sheet_data(df, sheet_name):
    """Processa os dados de uma aba com base na configuração definida."""
    normalized_sheet_name = normalize_sheet_name(sheet_name)
    actual_sheet_name = SHEET_NAME_MAPPING.get(normalized_sheet_name, sheet_name)
    
    if actual_sheet_name not in SHEET_CONFIG:
        st.warning(f"Aba '{actual_sheet_name}' não configurada em SHEET_CONFIG. Usando processamento genérico.")
        for col in df.columns:
            if df[col].dtype == 'datetime64[ns]':
                df[col] = df[col].dt.strftime('%d/%m/%Y %H:%M:%S')
            df[col] = df[col].apply(lambda x: str(x) if pd.notnull(x) else '').replace('nan', '').str.replace('\n', ' ', regex=False).str.strip()
        return df
    
    config = SHEET_CONFIG[actual_sheet_name]['columns']
    for col, col_config in config.items():
        if col not in df.columns:
            col_type = col_config['type']
            if col_type in ['date', 'datetime']:
                df[col] = pd.NaT
            elif col_type == 'boolean':
                df[col] = False
            elif col_type in ['int', 'float']:
                df[col] = 0
            elif col_type == 'categorical':
                df[col] = col_config.get('values', [''])[0]
            else:
                df[col] = ''
            continue
        
        col_type = col_config['type']
        if col_type == 'string':
            df[col] = df[col].astype(str).replace('nan', '').replace('-', '').str.replace('\n', ' ', regex=False).str.strip()
        elif col_type == 'categorical':
            allowed_values = col_config.get('values', [])
            df[col] = df[col].apply(lambda x: x if pd.notnull(x) and str(x).strip() in allowed_values else allowed_values[0] if allowed_values else '')
        elif col_type == 'date':
            df[col] = pd.to_datetime(df[col], format=col_config.get('format', '%d/%m/%Y'), errors='coerce')
            df[col] = df[col].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else '')
        elif col_type == 'datetime':
            df[col] = pd.to_datetime(df[col], format=col_config.get('format', '%d/%m/%Y %H:%M'), errors='coerce')
            df[col] = df[col].apply(lambda x: x.strftime('%d/%m/%Y %H:%M') if pd.notna(x) else '')
        elif col_type == 'boolean':
            df[col] = df[col].apply(lambda x: True if str(x).strip().upper() in ['X', 'SIM', 'TRUE', 'S'] else False)
        elif col_type == 'int':
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        elif col_type == 'float':
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
        elif col_type == 'training_period':
            df[[f'{col}_INÍCIO', f'{col}_FIM']] = df[col].apply(parse_training_period).apply(pd.Series)
            df.drop(columns=[col], inplace=True)
        elif col_type == 'phone':
            df[col] = df[col].apply(clean_phone_number)
            if col_config.get('allow_empty', False):
                df[col] = df[col].replace('', np.nan)
        elif col_type == 'email':
            df[col] = df[col].apply(clean_email)
        elif col_type == 'time':
            df[col] = df[col].apply(clean_time)
    
    return df

@st.cache_data
def process_excel_file(uploaded_file=None):
    """Processa todas as abas do arquivo Excel e retorna um dicionário de DataFrames."""
    processed_data = {}
    file_path = uploaded_file if uploaded_file is not None else EXCEL_FILE
    
    try:
        if isinstance(file_path, str) and not os.path.exists(file_path):
            st.warning(f"Arquivo não encontrado no caminho principal: {os.path.abspath(file_path)}")
            file_path = FALLBACK_EXCEL_FILE
            if not os.path.exists(file_path):
                st.error(f"Arquivo não encontrado no caminho alternativo: {os.path.abspath(file_path)}")
                st.markdown("""
                ### Possíveis Soluções
                - Verifique se o arquivo `ACOMPANHAMENTO_CIN_EM_TODO_LUGAR.xlsx` está em `C:\\Users\\re049227\\Documents\\python_graphs_CIN\\`.
                - Confirme se o nome do arquivo está correto (sem espaços extras ou caracteres ocultos).
                - Se o arquivo foi carregado via upload, certifique-se de que o módulo `upload_excel.py` está configurado corretamente.
                - Em produção, verifique o caminho do arquivo no servidor ou contêiner.
                - Certifique-se de que o arquivo não está corrompido ou bloqueado por outro programa.
                """)
                return processed_data
        
        xls = pd.ExcelFile(file_path, engine='openpyxl')
        st.write(f"[DEBUG] Abas disponíveis no arquivo Excel: {xls.sheet_names}")
        for sheet_name in SHEET_CONFIG.keys():
            normalized_sheet_name = normalize_sheet_name(sheet_name)
            actual_sheet_name = SHEET_NAME_MAPPING.get(normalized_sheet_name, sheet_name)
            if actual_sheet_name not in xls.sheet_names:
                st.warning(f"Aba '{actual_sheet_name}' não encontrada no arquivo Excel. Abas disponíveis: {xls.sheet_names}")
                processed_data[actual_sheet_name] = pd.DataFrame()
                continue
            try:
                df = load_excel(sheet_name, file_path)
                processed_data[actual_sheet_name] = process_sheet_data(df, sheet_name)
            except Exception as e:
                st.warning(f"Erro ao processar a aba {actual_sheet_name}: {str(e)}")
                processed_data[actual_sheet_name] = pd.DataFrame()
    except Exception as e:
        st.error(f"Erro ao abrir o arquivo Excel: {str(e)}")
        return processed_data
    
    return processed_data

def save_excel(df, sheet_name, file_path=EXCEL_FILE):
    """Salva um DataFrame em uma aba específica do arquivo Excel."""
    try:
        normalized_sheet_name = normalize_sheet_name(sheet_name)
        actual_sheet_name = SHEET_NAME_MAPPING.get(normalized_sheet_name, sheet_name)
        
        # Handle uploaded file
        if isinstance(file_path, BytesIO):
            wb = openpyxl.load_workbook(file_path)
        else:
            # Check if file exists
            if not os.path.exists(file_path):
                st.warning(f"Arquivo não encontrado no caminho principal: {os.path.abspath(file_path)}")
                file_path = FALLBACK_EXCEL_FILE
                if not os.path.exists(file_path):
                    wb = openpyxl.Workbook()
                    wb.save(file_path)
                    st.write(f"[DEBUG] Novo arquivo Excel criado em: {os.path.abspath(file_path)}")
            wb = openpyxl.load_workbook(file_path)
        
        # Verify or create sheet
        if actual_sheet_name not in wb.sheetnames:
            wb.create_sheet(actual_sheet_name)
            st.write(f"[DEBUG] Aba '{actual_sheet_name}' criada no arquivo Excel")
        
        # Select the sheet
        ws = wb[actual_sheet_name]
        
        # Clear existing sheet content
        ws.delete_rows(1, ws.max_row)
        
        # Add headers
        headers = df.columns.tolist()
        ws.append(headers)
        
        # Add data
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append([str(cell).replace('NaT', '') if pd.notna(cell) else '' for cell in row])
        
        # Adjust column widths
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        # Save the file
        if isinstance(file_path, BytesIO):
            output = BytesIO()
            wb.save(output)
            file_path.seek(0)
            file_path.write(output.getvalue())
        else:
            wb.save(file_path)
            st.write(f"[DEBUG] Arquivo salvo em: {os.path.abspath(file_path)}")
        
        st.success(f"Dados salvos com sucesso na aba '{actual_sheet_name}'!")
        return True
    except Exception as e:
        st.error(f"Erro ao salvar a aba {actual_sheet_name}: {str(e)}")
        return False